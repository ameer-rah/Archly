from __future__ import annotations

import os
import time
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from urllib.parse import urlparse

import requests
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


load_dotenv()

GOOGLE_MAPS_API_KEY = os.getenv("GOOGLE_MAPS_API_KEY", "").strip()
HUNTER_API_KEY = os.getenv("HUNTER_API_KEY", "").strip()

OUTPUT_FILE = os.getenv(
    "OUTPUT_FILE",
    "nyc_architecture_firms.xlsx",
)

MAX_PAGES_PER_QUERY = int(
    os.getenv("MAX_PAGES_PER_QUERY", "3")
)

ALLOW_PERSONAL_EMAILS = (
    os.getenv("ALLOW_PERSONAL_EMAILS", "false").lower() == "true"
)

PLACES_SEARCH_URL = (
    "https://places.googleapis.com/v1/places:searchText"
)

HUNTER_DOMAIN_SEARCH_URL = (
    "https://api.hunter.io/v2/domain-search"
)

# Google documents these bounds as a rectangle fully enclosing NYC.
NYC_RECTANGLE = {
    "low": {
        "latitude": 40.477398,
        "longitude": -74.259087,
    },
    "high": {
        "latitude": 40.91618,
        "longitude": -73.70018,
    },
}

# Searching by borough generally provides better coverage than one broad query.
SEARCH_QUERIES = [
    "architecture firms in Manhattan, New York",
    "architecture firms in Brooklyn, New York",
    "architecture firms in Queens, New York",
    "architecture firms in the Bronx, New York",
    "architecture firms in Staten Island, New York",
]

PLACES_FIELD_MASK = ",".join(
    [
        "places.id",
        "places.displayName",
        "places.formattedAddress",
        "places.websiteUri",
        "places.googleMapsUri",
        "places.businessStatus",
        "nextPageToken",
    ]
)

session = requests.Session()
session.headers.update(
    {
        "User-Agent": (
            "NYCArchitectureFirmResearch/1.0 "
            "(business-directory-research)"
        )
    }
)


def require_environment_variables() -> None:
    """Stop immediately when required configuration is missing."""
    missing: list[str] = []

    if not GOOGLE_MAPS_API_KEY:
        missing.append("GOOGLE_MAPS_API_KEY")

    if not HUNTER_API_KEY:
        missing.append("HUNTER_API_KEY")

    if missing:
        joined = ", ".join(missing)
        raise RuntimeError(
            f"Missing required environment variable(s): {joined}"
        )


def search_google_places(query: str) -> list[dict[str, Any]]:
    """
    Search Google Places for one query and follow pagination.

    All request parameters other than pageToken remain unchanged
    between pagination calls.
    """
    headers = {
        "Content-Type": "application/json",
        "X-Goog-Api-Key": GOOGLE_MAPS_API_KEY,
        "X-Goog-FieldMask": PLACES_FIELD_MASK,
    }

    base_body: dict[str, Any] = {
        "textQuery": query,
        "pageSize": 20,
        "languageCode": "en",
        "regionCode": "US",
        "locationRestriction": {
            "rectangle": NYC_RECTANGLE,
        },
    }

    results: list[dict[str, Any]] = []
    page_token: str | None = None

    for page_number in range(1, MAX_PAGES_PER_QUERY + 1):
        body = dict(base_body)

        if page_token:
            body["pageToken"] = page_token

        print(
            f"Google Places: {query!r}, page {page_number}"
        )

        response = session.post(
            PLACES_SEARCH_URL,
            headers=headers,
            json=body,
            timeout=30,
        )
        response.raise_for_status()

        payload = response.json()
        results.extend(payload.get("places", []))

        page_token = payload.get("nextPageToken")

        if not page_token:
            break

        # Modest delay between pagination requests.
        time.sleep(1)

    return results


def extract_domain(website_url: str) -> str:
    """Convert a full website URL into a clean domain."""
    if not website_url:
        return ""

    value = website_url.strip()

    if not value.startswith(("http://", "https://")):
        value = f"https://{value}"

    parsed = urlparse(value)
    domain = parsed.netloc.lower()

    if domain.startswith("www."):
        domain = domain[4:]

    return domain.split(":")[0]


def choose_hunter_email(
    emails: list[dict[str, Any]],
) -> dict[str, Any] | None:
    """
    Prefer generic business addresses such as info@ and contact@.

    Personal employee emails are excluded unless
    ALLOW_PERSONAL_EMAILS=true.
    """
    valid_emails = [
        email
        for email in emails
        if email.get("value")
    ]

    if not ALLOW_PERSONAL_EMAILS:
        valid_emails = [
            email
            for email in valid_emails
            if email.get("type") == "generic"
        ]

    if not valid_emails:
        return None

    preferred_local_parts = [
        "info",
        "contact",
        "hello",
        "office",
        "studio",
        "inquiries",
        "admin",
        "mail",
        "general",
    ]

    def ranking(email: dict[str, Any]) -> tuple[int, int, int]:
        address = str(email.get("value", "")).lower()
        local_part = address.split("@")[0]

        generic_rank = (
            0 if email.get("type") == "generic" else 1
        )

        try:
            name_rank = preferred_local_parts.index(local_part)
        except ValueError:
            name_rank = len(preferred_local_parts)

        confidence = int(email.get("confidence") or 0)

        # Lower tuple sorts first. Negative confidence means higher
        # confidence is preferred.
        return generic_rank, name_rank, -confidence

    return sorted(valid_emails, key=ranking)[0]


def find_email_with_hunter(
    domain: str,
) -> dict[str, Any]:
    """Use Hunter Domain Search to find a business email."""
    blank_result = {
        "email": "",
        "email_type": "",
        "confidence": "",
        "email_source": "",
    }

    if not domain:
        return blank_result

    try:
        response = session.get(
            HUNTER_DOMAIN_SEARCH_URL,
            headers={
                "X-API-KEY": HUNTER_API_KEY,
            },
            params={
                "domain": domain,
                "limit": 10,
            },
            timeout=30,
        )

        if response.status_code == 404:
            return blank_result

        response.raise_for_status()

        payload = response.json()
        emails = payload.get("data", {}).get("emails", [])

        selected = choose_hunter_email(emails)

        if not selected:
            return blank_result

        sources = selected.get("sources", [])
        first_source = ""

        if sources:
            first_source = (
                sources[0].get("uri")
                or sources[0].get("domain")
                or ""
            )

        return {
            "email": selected.get("value", ""),
            "email_type": selected.get("type", ""),
            "confidence": selected.get("confidence", ""),
            "email_source": first_source,
        }

    except requests.RequestException as error:
        print(f"Hunter error for {domain}: {error}")
        return blank_result


def collect_firms() -> list[dict[str, Any]]:
    """Collect, deduplicate, and enrich architecture firms."""
    firms_by_place_id: dict[str, dict[str, Any]] = {}

    for query in SEARCH_QUERIES:
        try:
            places = search_google_places(query)
        except requests.RequestException as error:
            print(f"Google Places error for {query}: {error}")
            continue

        for place in places:
            place_id = place.get("id", "")

            if not place_id:
                continue

            business_status = place.get("businessStatus", "")

            if business_status == "CLOSED_PERMANENTLY":
                continue

            existing = firms_by_place_id.get(place_id)

            if existing:
                existing_queries = existing["search_queries"]

                if query not in existing_queries:
                    existing_queries.append(query)

                continue

            display_name = place.get("displayName", {})
            website = place.get("websiteUri", "")
            domain = extract_domain(website)

            firms_by_place_id[place_id] = {
                "name": display_name.get("text", ""),
                "address": place.get("formattedAddress", ""),
                "website": website,
                "domain": domain,
                "business_status": business_status,
                "place_id": place_id,
                "google_maps_url": place.get(
                    "googleMapsUri",
                    "",
                ),
                "search_queries": [query],
            }

    firms = list(firms_by_place_id.values())
    firms.sort(key=lambda firm: firm["name"].lower())

    total = len(firms)

    for index, firm in enumerate(firms, start=1):
        domain = firm["domain"]

        print(
            f"Hunter {index}/{total}: "
            f"{firm['name']} ({domain or 'no website'})"
        )

        email_result = find_email_with_hunter(domain)
        firm.update(email_result)

        # Avoid sending all provider requests at once.
        time.sleep(0.15)

    return firms


def save_to_excel(
    firms: list[dict[str, Any]],
    output_path: str,
) -> None:
    """Write the completed dataset to a formatted Excel workbook."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "NYC Architecture Firms"

    headers = [
        "Firm Name",
        "Address",
        "Website",
        "Domain",
        "Email",
        "Email Type",
        "Email Confidence",
        "Email Source",
        "Business Status",
        "Google Place ID",
        "Google Maps URL",
        "Search Query",
        "Collected At",
    ]

    worksheet.append(headers)

    collected_at = datetime.now(timezone.utc).isoformat(
        timespec="seconds"
    )

    for firm in firms:
        worksheet.append(
            [
                firm["name"],
                firm["address"],
                firm["website"],
                firm["domain"],
                firm["email"],
                firm["email_type"],
                firm["confidence"],
                firm["email_source"],
                firm["business_status"],
                firm["place_id"],
                firm["google_maps_url"],
                "; ".join(firm["search_queries"]),
                collected_at,
            ]
        )

        row_number = worksheet.max_row

        hyperlink_columns = {
            3: firm["website"],
            8: firm["email_source"],
            11: firm["google_maps_url"],
        }

        for column_number, url in hyperlink_columns.items():
            if url:
                cell = worksheet.cell(
                    row=row_number,
                    column=column_number,
                )
                cell.hyperlink = url
                cell.style = "Hyperlink"

        email = firm["email"]

        if email:
            email_cell = worksheet.cell(
                row=row_number,
                column=5,
            )
            email_cell.hyperlink = f"mailto:{email}"
            email_cell.style = "Hyperlink"

    # Header styling
    for cell in worksheet[1]:
        cell.font = Font(bold=True)

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    column_widths = {
        1: 34,
        2: 45,
        3: 40,
        4: 30,
        5: 34,
        6: 14,
        7: 18,
        8: 40,
        9: 20,
        10: 34,
        11: 40,
        12: 42,
        13: 26,
    }

    for column_number, width in column_widths.items():
        column_letter = get_column_letter(column_number)
        worksheet.column_dimensions[column_letter].width = width

    if worksheet.max_row > 1:
        table_reference = (
            f"A1:M{worksheet.max_row}"
        )

        table = Table(
            displayName="NYCArchitectureFirms",
            ref=table_reference,
        )

        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )

        worksheet.add_table(table)

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)

    workbook.save(output)

    print(
        f"\nSaved {len(firms)} firms to:\n"
        f"{output.resolve()}"
    )


def main() -> None:
    require_environment_variables()
    firms = collect_firms()
    save_to_excel(firms, OUTPUT_FILE)


if __name__ == "__main__":
    main()